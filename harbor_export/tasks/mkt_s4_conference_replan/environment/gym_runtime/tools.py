"""
Tool layer — schemas and executor for agent tool-use.

Exposes a compact set of tools that collectively cover ~80% of scenario action
surface across all 25 scenarios: messaging, email, calendar, documents,
approvals, file access, and session termination. Each tool execution:

1. Constructs a `gym_core.Action` with the right fields
2. Appends it to the ActionRecorder
3. Advances the SimClock
4. Optionally invokes the NPCSimulator to generate NPC responses
5. Returns a string tool result the agent sees in its next turn

Adding a new tool: define its schema in TOOL_SCHEMAS and a handler in
ToolExecutor. Each handler returns the tool result string.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import Any, TYPE_CHECKING

from gym_core import Action, ActionType, Severity
from gym_runtime.clock import SimClock
from gym_runtime.inbox import Inbox, InboxItem
from gym_runtime.recorder import ActionRecorder

if TYPE_CHECKING:
    from gym_runtime.scenario import Scenario
    from gym_runtime.npc import NPCSimulator


# ---------------------------------------------------------------------------
# Tool schemas (Anthropic tool-use format — Claude / Gemini compatible)
# ---------------------------------------------------------------------------

TOOL_SCHEMAS: list[dict[str, Any]] = [
    {
        "name": "send_slack_message",
        "description": (
            "Send a Slack DM or channel message. Use 'dm' for direct messages. "
            "For channel messages set channel to e.g. '#deal-brightloom'. "
            "Recipient is a persona_id for DMs, or the channel name for channel posts."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "to": {"type": "string", "description": "persona_id (for DM) or channel name"},
                "channel": {"type": "string", "enum": ["dm", "channel"], "description": "dm or channel"},
                "body": {"type": "string"},
                "thread_id": {"type": "string", "description": "optional reply thread id"},
            },
            "required": ["to", "channel", "body"],
        },
    },
    {
        "name": "send_email",
        "description": (
            "Send an email. Use for cross-functional requests, formal approvals, or "
            "recipients who prefer email. Use encrypted=true for legal, HR privileged, "
            "security, or financial disclosure topics."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "to": {"type": "array", "items": {"type": "string"}, "description": "persona_ids or external email strings"},
                "cc": {"type": "array", "items": {"type": "string"}},
                "subject": {"type": "string"},
                "body": {"type": "string"},
                "encrypted": {"type": "boolean", "default": False},
                "classification": {"type": "string", "enum": ["public", "internal", "confidential", "restricted", "privileged"]},
                "attachments": {"type": "array", "items": {"type": "string"}, "description": "paths to inputs/ files or named artifacts"},
            },
            "required": ["to", "subject", "body"],
        },
    },
    {
        "name": "create_calendar_invite",
        "description": (
            "Create a calendar invite. Attendees are persona_ids. Time must be ISO-8601 "
            "with timezone. Check each attendee's working hours before scheduling."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "attendees": {"type": "array", "items": {"type": "string"}},
                "subject": {"type": "string"},
                "start_time": {"type": "string", "description": "ISO-8601 timestamp with tz"},
                "duration_minutes": {"type": "integer"},
                "body": {"type": "string", "description": "agenda / context"},
                "location": {"type": "string"},
            },
            "required": ["attendees", "subject", "start_time", "duration_minutes"],
        },
    },
    {
        "name": "create_document",
        "description": (
            "Create a new document (Google Doc / Notion page / shared memo). Use for "
            "deliverables like reconciliation memos, briefing docs, policy drafts, "
            "status reports, decision memos."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "title": {"type": "string"},
                "body": {"type": "string"},
                "shared_with": {"type": "array", "items": {"type": "string"}, "description": "persona_ids"},
                "classification": {"type": "string"},
            },
            "required": ["title", "body"],
        },
    },
    {
        "name": "request_approval",
        "description": (
            "Request an explicit approval from a named approver (e.g. Deal Desk "
            "discount, pricing exception, publication sign-off, SOX remediation claim). "
            "Returns pending until the approver responds in inbox."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "approver": {"type": "string", "description": "persona_id"},
                "subject": {"type": "string"},
                "body": {"type": "string"},
                "amount": {"type": "number"},
                "currency": {"type": "string"},
            },
            "required": ["approver", "subject", "body"],
        },
    },
    {
        "name": "docusign_envelope",
        "description": (
            "Create a DocuSign envelope for signature. Use for contracts, separation "
            "agreements, claims review workflows, and any legal artifact that must "
            "travel through a signature trail."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "signers": {"type": "array", "items": {"type": "string"}, "description": "persona_ids in signing order"},
                "subject": {"type": "string"},
                "body": {"type": "string"},
                "document_reference": {"type": "string"},
            },
            "required": ["signers", "subject"],
        },
    },
    {
        "name": "list_files",
        "description": (
            "List all files in the scenario's inputs/ directory. Every scenario starts "
            "with pre-populated context the agent must read (emails, chat logs, "
            "spreadsheets, reports, policy docs). Call this early to discover them."
        ),
        "input_schema": {"type": "object", "properties": {}},
    },
    {
        "name": "read_file",
        "description": (
            "Read a file from the scenario's inputs/ directory. Path is relative to "
            "inputs/, e.g. 'emails/01_ayo_kickoff.md' or 'sheets/merit_pool_v1.xlsx'. "
            "Binary files (.xlsx, .pdf) are returned as a summary; text files are "
            "returned verbatim."
        ),
        "input_schema": {
            "type": "object",
            "properties": {"path": {"type": "string"}},
            "required": ["path"],
        },
    },
    {
        "name": "check_inbox",
        "description": (
            "Check for new unread messages — Slack DMs, emails, calendar responses, "
            "approval decisions. NPCs may respond asynchronously to your actions; "
            "call this periodically to see their replies. Advances sim time slightly."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "channel": {"type": "string", "description": "optional: 'slack', 'email', or omit for all"},
            },
        },
    },
    {
        "name": "finish_scenario",
        "description": (
            "Declare the scenario complete. Call this only when you believe every "
            "required deliverable has been produced and every dependency satisfied. "
            "Ends the session and triggers scoring."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "summary": {"type": "string", "description": "brief recap of what you accomplished"},
            },
            "required": ["summary"],
        },
    },
]


# ---------------------------------------------------------------------------
# Executor
# ---------------------------------------------------------------------------

@dataclass
class ToolExecutionResult:
    tool_name: str
    ok: bool
    result: str
    advance_minutes: float


class ToolExecutor:
    """Executes a tool call: logs an Action, advances the clock, optionally
    triggers NPC reactions, returns a tool-result string for the agent."""

    def __init__(
        self,
        scenario: "Scenario",
        recorder: ActionRecorder,
        clock: SimClock,
        inbox: Inbox,
        npc: "NPCSimulator | None" = None,
    ):
        self.scenario = scenario
        self.recorder = recorder
        self.clock = clock
        self.inbox = inbox
        self.npc = npc
        self.finished = False
        self.finish_summary: str | None = None

    # ------------------------------------------------------------------

    def execute(self, name: str, tool_input: dict[str, Any]) -> ToolExecutionResult:
        handler = getattr(self, f"_tool_{name}", None)
        if handler is None:
            return ToolExecutionResult(
                tool_name=name, ok=False,
                result=f"Unknown tool: {name}. Available: {[t['name'] for t in TOOL_SCHEMAS]}",
                advance_minutes=0,
            )
        try:
            result_str, advance = handler(tool_input)
            self.clock.advance(advance)
            return ToolExecutionResult(name, True, result_str, advance)
        except Exception as exc:
            return ToolExecutionResult(name, False, f"ERROR: {type(exc).__name__}: {exc}", 0)

    # ------------------------------------------------------------------
    # Handlers
    # ------------------------------------------------------------------

    def _tool_send_slack_message(self, inp: dict) -> tuple[str, float]:
        to = inp["to"]
        channel = inp["channel"]
        body = inp["body"]
        action = self.recorder.build_and_record(
            action_type=ActionType.SLACK_MESSAGE,
            timestamp=self.clock.now,
            tool="slack",
            recipient=to,
            channel="dm" if channel == "dm" else to,
            body=body,
            metadata={"thread_id": inp.get("thread_id")},
        )
        if self.npc:
            self.npc.react_to_action(action, self.inbox)
        return (f"Slack message sent to {to} at {self.clock.iso()}.", 3)

    def _tool_send_email(self, inp: dict) -> tuple[str, float]:
        encrypted = bool(inp.get("encrypted", False))
        atype = ActionType.ENCRYPTED_EMAIL if encrypted else ActionType.EMAIL
        action = self.recorder.build_and_record(
            action_type=atype,
            timestamp=self.clock.now,
            tool="outlook_encrypted" if encrypted else "email",
            recipient=(inp["to"][0] if inp["to"] else None),
            cc=inp.get("cc", [])[:],
            subject=inp.get("subject"),
            body=inp.get("body"),
            classification=inp.get("classification"),
            attachments=inp.get("attachments", [])[:],
            channel="encrypted_outlook" if encrypted else "email",
            metadata={"all_recipients": inp["to"]},
        )
        # For multi-recipient emails, record a follow-up Action per additional recipient so
        # verifiers that filter by recipient see each of them.
        for extra in inp["to"][1:]:
            self.recorder.build_and_record(
                action_type=atype,
                timestamp=self.clock.now,
                tool="outlook_encrypted" if encrypted else "email",
                recipient=extra,
                cc=inp.get("cc", [])[:],
                subject=inp.get("subject"),
                body=inp.get("body"),
                classification=inp.get("classification"),
                attachments=inp.get("attachments", [])[:],
                channel="encrypted_outlook" if encrypted else "email",
                metadata={"all_recipients": inp["to"], "multi_recipient_echo": True},
            )
        if self.npc:
            self.npc.react_to_action(action, self.inbox)
        return (
            f"Email sent to {len(inp['to'])} recipient(s){' (encrypted)' if encrypted else ''}.",
            5,
        )

    def _tool_create_calendar_invite(self, inp: dict) -> tuple[str, float]:
        start_time = self._parse_iso(inp["start_time"])
        duration = int(inp["duration_minutes"])
        end_time = start_time + timedelta(minutes=duration) if start_time else None
        action = self.recorder.build_and_record(
            action_type=ActionType.CALENDAR_INVITE,
            timestamp=self.clock.now,
            tool="calendar",
            subject=inp["subject"],
            body=inp.get("body"),
            attendees=list(inp["attendees"]),
            start_time=start_time,
            end_time=end_time,
            duration_minutes=duration,
            metadata={"location": inp.get("location")},
        )
        if self.npc:
            self.npc.react_to_action(action, self.inbox)
        return (
            f"Calendar invite created for {inp['subject']} at {inp['start_time']} "
            f"({duration} min) with {len(inp['attendees'])} attendee(s).",
            3,
        )

    def _tool_create_document(self, inp: dict) -> tuple[str, float]:
        action = self.recorder.build_and_record(
            action_type=ActionType.DOCUMENT_CREATE,
            timestamp=self.clock.now,
            tool="docs",
            subject=inp["title"],
            body=inp["body"],
            classification=inp.get("classification"),
            metadata={"shared_with": inp.get("shared_with", [])},
        )
        return (f"Document '{inp['title']}' created ({len(inp['body'])} chars).", 8)

    def _tool_request_approval(self, inp: dict) -> tuple[str, float]:
        action = self.recorder.build_and_record(
            action_type=ActionType.APPROVAL_REQUEST,
            timestamp=self.clock.now,
            tool="approval",
            recipient=inp["approver"],
            subject=inp["subject"],
            body=inp.get("body"),
            amount=inp.get("amount"),
            currency=inp.get("currency"),
            approval_status="pending",
        )
        if self.npc:
            self.npc.react_to_action(action, self.inbox)
        return (
            f"Approval request sent to {inp['approver']}. Status: pending.",
            3,
        )

    def _tool_docusign_envelope(self, inp: dict) -> tuple[str, float]:
        self.recorder.build_and_record(
            action_type=ActionType.DOCUSIGN_ENVELOPE,
            timestamp=self.clock.now,
            tool="docusign",
            subject=inp["subject"],
            body=inp.get("body"),
            attendees=list(inp["signers"]),
            metadata={"document_reference": inp.get("document_reference")},
        )
        return (f"DocuSign envelope created with {len(inp['signers'])} signer(s).", 4)

    def _tool_list_files(self, inp: dict) -> tuple[str, float]:
        files = self.scenario.list_inputs()
        if not files:
            return ("(inputs/ is empty)", 1)
        return ("Available input files:\n" + "\n".join(f"  inputs/{p}" for p in files), 1)

    def _tool_read_file(self, inp: dict) -> tuple[str, float]:
        path = inp["path"]
        # Normalize: accept "inputs/foo" or "foo"
        if path.startswith("inputs/"):
            path = path[len("inputs/"):]
        try:
            raw = self.scenario.read_input(path)
        except FileNotFoundError:
            return (f"File not found: inputs/{path}. Call list_files to see available files.", 1)

        # Stamp a FILE_DOWNLOAD action so verifiers that check audit-trail
        # evidence (e.g. "netsuite_export_downloaded") can see that the agent
        # pulled the source file. Reading a pre-staged input IS the download
        # from the scenario's perspective.
        self.recorder.build_and_record(
            action_type=ActionType.FILE_DOWNLOAD,
            timestamp=self.clock.now,
            tool="read_file",
            subject=f"inputs/{path}",
            body=None,
            metadata={"source_path": path, "size_bytes": len(raw)},
        )

        # Text vs binary handling
        lower = path.lower()
        if lower.endswith((".md", ".yaml", ".yml", ".txt", ".csv", ".json")):
            text = raw.decode("utf-8", errors="replace")
            if len(text) > 20000:
                text = text[:20000] + f"\n... (truncated, file has {len(raw)} bytes total)"
            return (text, 2)
        if lower.endswith(".xlsx"):
            return (_summarize_xlsx(raw, path), 3)
        if lower.endswith(".pdf"):
            return (_summarize_pdf(raw, path), 3)
        return (f"(binary file of {len(raw)} bytes at inputs/{path})", 1)

    def _tool_check_inbox(self, inp: dict) -> tuple[str, float]:
        channel = inp.get("channel") if inp else None
        items = self.inbox.unread_by_channel(channel) if channel else self.inbox.unread()
        for item in items:
            item.read = True
        if not items:
            return ("(no unread messages)", 1)
        return (
            f"{len(items)} unread message(s):\n\n" + self.inbox.snapshot(only_unread=False)
                if False else
            f"{len(items)} unread message(s):\n\n" + "\n\n---\n\n".join(i.format() for i in items),
            1,
        )

    def _tool_finish_scenario(self, inp: dict) -> tuple[str, float]:
        self.finished = True
        self.finish_summary = inp.get("summary", "")
        return (f"Scenario marked complete. Summary recorded: {self.finish_summary[:200]}", 0)

    # ------------------------------------------------------------------

    def _parse_iso(self, s: str) -> datetime | None:
        try:
            return datetime.fromisoformat(s)
        except (ValueError, TypeError):
            return None


def _summarize_xlsx(raw: bytes, path: str) -> str:
    try:
        import io
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(raw), data_only=True)
        lines = [f"[xlsx] inputs/{path}"]
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"\n=== Sheet: {sheet_name} ({ws.max_row} rows × {ws.max_column} cols) ===")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 50:
                    lines.append(f"... ({ws.max_row - 50} more rows)")
                    break
                lines.append(" | ".join("" if v is None else str(v) for v in row))
        return "\n".join(lines)
    except Exception as exc:
        return f"[xlsx] inputs/{path} ({len(raw)} bytes) — could not parse: {exc}"


def _summarize_pdf(raw: bytes, path: str) -> str:
    # Try pypdf first (lightweight), fall back to declaring we cannot render.
    try:
        import io
        try:
            from pypdf import PdfReader
        except ImportError:
            return (f"[pdf] inputs/{path} ({len(raw)} bytes) — install pypdf for content extraction. "
                    f"File name and scenario context usually imply its purpose.")
        reader = PdfReader(io.BytesIO(raw))
        text = "\n".join((p.extract_text() or "") for p in reader.pages[:5])
        if len(text) > 10000:
            text = text[:10000] + "\n... (truncated)"
        return f"[pdf] inputs/{path} ({len(reader.pages)} pages)\n\n{text}"
    except Exception as exc:
        return f"[pdf] inputs/{path} ({len(raw)} bytes) — parse failed: {exc}"
