#!/usr/bin/env python3
"""Build the Harbor visualizer by embedding task and trajectory data into the HTML."""

import json
import os
import re
from pathlib import Path

import yaml

import csv
import io

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import pypdf
except ImportError:
    pypdf = None

BASE = Path(__file__).parent / "harbor_export"
OUT = Path(__file__).parent / "index.html"
TEMPLATE = Path(__file__).parent / "visualizer.html"

# Domain folder name mapping from task id prefix
DOMAIN_MAP = {
    "cs": "customer_support",
    "fin": "finance",
    "hr": "hr",
    "mkt": "marketing",
    "sp": "sales_procurement",
}


def scenario_folder(task_id):
    """Derive the scenario subfolder name from the task id.
    E.g. cs_s1_churn_rescue -> s1_churn_rescue"""
    parts = task_id.split("_", 1)
    return parts[1] if len(parts) > 1 else task_id


def find_domain_dir(env_dir, task_id):
    """Find the domain directory under gym/domains/."""
    prefix = task_id.split("_")[0]
    domain = DOMAIN_MAP.get(prefix)
    if domain:
        d = env_dir / "gym" / "domains" / domain
        if d.exists():
            return d
    # Fallback: search
    gym_domains = env_dir / "gym" / "domains"
    if gym_domains.exists():
        for d in gym_domains.iterdir():
            if d.is_dir():
                return d
    return None


def load_input_file(fp):
    """Load a single input file and return a structured entry with type info."""
    suffix = fp.suffix.lower()

    # YAML (chat logs)
    if suffix in (".yaml", ".yml"):
        try:
            raw = fp.read_text()
            data = yaml.safe_load(raw)
            # Detect chat-style yaml (has 'messages' key)
            if isinstance(data, dict) and "messages" in data:
                return {"type": "chat", "data": data, "raw": raw}
            return {"type": "yaml", "raw": raw}
        except Exception:
            return {"type": "text", "raw": fp.read_text()}

    # Markdown — detect email (has YAML frontmatter with from/to/subject)
    if suffix == ".md":
        try:
            raw = fp.read_text()
            # Check for email-style frontmatter
            if raw.startswith("---"):
                parts = raw.split("---", 2)
                if len(parts) >= 3:
                    fm = yaml.safe_load(parts[1])
                    if isinstance(fm, dict) and ("from" in fm or "to" in fm or "subject" in fm):
                        body = parts[2].strip()
                        if len(body) > 5000:
                            body = body[:5000] + "\n...[truncated]"
                        return {"type": "email", "frontmatter": fm, "body": body}
            if len(raw) > 5000:
                raw = raw[:5000] + "\n...[truncated]"
            return {"type": "markdown", "raw": raw}
        except Exception:
            return {"type": "text", "raw": fp.read_text()[:5000]}

    # CSV
    if suffix == ".csv":
        try:
            raw = fp.read_text()
            reader = csv.reader(io.StringIO(raw))
            rows = list(reader)
            headers = rows[0] if rows else []
            data_rows = rows[1:50]  # cap at 50 rows
            return {"type": "csv", "headers": headers, "rows": data_rows}
        except Exception:
            return {"type": "text", "raw": fp.read_text()[:5000]}

    # Excel
    if suffix == ".xlsx" and openpyxl:
        try:
            wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
            sheets = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(max_row=60, values_only=True):
                    rows.append([str(c) if c is not None else "" for c in row])
                if rows:
                    sheets[sheet_name] = {"headers": rows[0], "rows": rows[1:]}
            wb.close()
            return {"type": "excel", "sheets": sheets}
        except Exception as e:
            return {"type": "binary", "info": f".xlsx file — {fp.stat().st_size} bytes (error: {e})"}

    # PDF
    if suffix == ".pdf" and pypdf:
        try:
            reader = pypdf.PdfReader(fp)
            pages = []
            for i, page in enumerate(reader.pages[:10]):
                text = page.extract_text() or ""
                if len(text) > 3000:
                    text = text[:3000] + "\n...[truncated]"
                pages.append({"page": i + 1, "text": text})
            return {"type": "pdf", "pages": pages, "total_pages": len(reader.pages)}
        except Exception as e:
            return {"type": "binary", "info": f".pdf file — {fp.stat().st_size} bytes (error: {e})"}

    # Other text files
    if suffix in (".txt", ".json", ".toml"):
        try:
            raw = fp.read_text()
            if len(raw) > 5000:
                raw = raw[:5000] + "\n...[truncated]"
            return {"type": "text", "raw": raw}
        except Exception:
            pass

    return {"type": "binary", "info": f"{suffix} file — {fp.stat().st_size} bytes"}


def load_tasks():
    tasks = {}
    tasks_dir = BASE / "tasks"
    for task_dir in sorted(tasks_dir.iterdir()):
        if not task_dir.is_dir():
            continue
        toml_path = task_dir / "task.toml"
        if not toml_path.exists():
            continue
        task = {"id": task_dir.name}
        current_section = None
        with open(toml_path) as f:
            for line in f:
                line = line.strip()
                if line.startswith("[") and line.endswith("]"):
                    current_section = line[1:-1]
                    continue
                if "=" not in line or line.startswith("#"):
                    continue
                key, val = line.split("=", 1)
                key = key.strip()
                val = val.strip().strip('"')
                # Unescape \n in TOML string values
                if isinstance(val, str):
                    val = val.replace('\\n', '\n')
                try:
                    val = int(val)
                except ValueError:
                    try:
                        val = float(val)
                    except ValueError:
                        pass
                if current_section in ("metadata", "task"):
                    task[key] = val

        # Load environment data
        env_dir = task_dir / "environment"
        task_id = task_dir.name

        # SKILLS.md
        skills_path = env_dir / "SKILLS.md"
        if skills_path.exists():
            task["skills_md"] = skills_path.read_text()

        # Find domain directory
        domain_dir = find_domain_dir(env_dir, task_id)
        if domain_dir:
            # Personas
            personas_path = domain_dir / "_personas.yaml"
            if personas_path.exists():
                pdata = yaml.safe_load(personas_path.read_text())
                if isinstance(pdata, dict):
                    task["organization"] = pdata.get("organization", {})
                    personas_list = pdata.get("personas", [])
                else:
                    personas_list = pdata if isinstance(pdata, list) else []
                # Extract key fields for each persona (trim for size)
                trimmed_personas = []
                for p in personas_list:
                    tp = {
                        "id": p.get("id", ""),
                        "name": p.get("name", ""),
                        "role": p.get("role", ""),
                        "reports_to": p.get("reports_to", ""),
                        "location": p.get("location", ""),
                        "timezone": p.get("timezone", ""),
                        "age": p.get("age"),
                    }
                    # Availability
                    avail = p.get("availability", {})
                    if avail:
                        tp["availability"] = {
                            "working_hours": avail.get("working_hours", ""),
                            "slack_responsiveness": avail.get("slack_responsiveness", ""),
                        }
                    # Communication style summary
                    comm = p.get("communication_style", {})
                    if comm:
                        tp["communication_style"] = {
                            k: v for k, v in comm.items()
                            if isinstance(v, str)
                        }
                        # Include pet_peeves if present
                        if "pet_peeves" in comm:
                            tp["communication_style"]["pet_peeves"] = comm["pet_peeves"]
                    # Decision making
                    dm = p.get("decision_making", {})
                    if dm:
                        tp["decision_making"] = dm
                    # KPIs
                    if "kpis" in p:
                        tp["kpis"] = p["kpis"]
                    # Relationships
                    if "relationships" in p:
                        tp["relationships"] = p["relationships"]
                    # Preferred/banned channels
                    if "preferred_channels" in p:
                        tp["preferred_channels"] = p["preferred_channels"]
                    if "banned_channels" in p:
                        tp["banned_channels"] = p["banned_channels"]
                    # Pain points, blind spots, triggers
                    for field in ("pain_points", "blind_spots", "triggers"):
                        if field in p:
                            tp[field] = p[field]
                    trimmed_personas.append(tp)
                task["personas"] = trimmed_personas

            # DAG
            scenario_dir = domain_dir / scenario_folder(task_id)
            dag_path = scenario_dir / "dag.md"
            if dag_path.exists():
                task["dag_md"] = dag_path.read_text()

            # Rubrics
            rubrics_path = scenario_dir / "rubrics.yaml"
            if rubrics_path.exists():
                rdata = yaml.safe_load(rubrics_path.read_text())
                task["rubrics"] = rdata

            # Input file listing — structured by type
            inputs_dir = scenario_dir / "inputs"
            if inputs_dir.exists():
                input_files = {}
                for fp in sorted(inputs_dir.rglob("*")):
                    if fp.is_file():
                        rel = str(fp.relative_to(inputs_dir))
                        entry = load_input_file(fp)
                        input_files[rel] = entry
                task["input_files"] = input_files

        tasks[task_dir.name] = task
    return tasks


def load_trajectories():
    trajs = {}
    traj_dir = BASE / "trajectories"
    for f in sorted(traj_dir.iterdir()):
        if not f.name.endswith(".atif.json"):
            continue
        with open(f) as fh:
            data = json.load(fh)

        steps = data.get("steps", [])
        trimmed_steps = []
        for step in steps:
            s = {k: v for k, v in step.items() if k != "observation"}
            if "observation" in step:
                obs = step["observation"]
                trimmed_results = []
                for r in obs.get("results", []):
                    content = r.get("content", "")
                    if isinstance(content, str) and len(content) > 1500:
                        content = content[:1500] + "\n...[truncated]"
                    trimmed_results.append({
                        "source_call_id": r.get("source_call_id"),
                        "content": content,
                        "is_error": r.get("is_error", False),
                    })
                s["observation"] = {"results": trimmed_results}
            if "message" in s and isinstance(s["message"], str) and len(s["message"]) > 3000:
                s["message"] = s["message"][:3000] + "\n...[truncated]"
            trimmed_steps.append(s)

        trajs[f.name] = {
            "agent": data.get("agent"),
            "final_metrics": data.get("final_metrics"),
            "steps": trimmed_steps,
        }
    return trajs


def build():
    print("Loading tasks...")
    tasks = load_tasks()
    print(f"  Found {len(tasks)} tasks")

    print("Loading trajectories...")
    trajs = load_trajectories()
    print(f"  Found {len(trajs)} trajectories")

    print("Reading template...")
    html = TEMPLATE.read_text()

    data_script = f"<script>window.__TASKS__ = {json.dumps(tasks, default=str)}; window.__TRAJECTORIES__ = {json.dumps(trajs, default=str)};</script>"
    html = html.replace(
        '<script>window.__TASKS__ = {}; window.__TRAJECTORIES__ = {};</script>',
        data_script,
    )

    OUT.write_text(html)
    size_mb = OUT.stat().st_size / 1024 / 1024
    print(f"Built {OUT} ({size_mb:.1f} MB)")


if __name__ == "__main__":
    build()
